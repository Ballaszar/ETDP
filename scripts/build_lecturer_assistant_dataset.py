from datetime import datetime, timezone
from pathlib import Path
import json
import re

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / 'Requests' / 'Developers-KnowledgeBase-4.0.xlsx'
OUTPUT_PATH = ROOT / 'frontend' / 'public' / 'data' / 'lecturer-assistant-links.json'


def clean(value):
    if value is None:
        return ''
    return str(value).replace('\r\n', '\n').replace('\r', '\n').strip()


def split_urls(value):
    raw = clean(value)
    if not raw:
        return []

    tokens = re.split(r'[\s,;]+', raw)
    urls = []
    for token in tokens:
        candidate = token.strip().strip(').]"\'')
        if candidate.lower().startswith('http://') or candidate.lower().startswith('https://'):
            urls.append(candidate)

    seen = set()
    deduped = []
    for url in urls:
        key = url.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(url)
    return deduped


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f'Workbook not found: {XLSX_PATH}')

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Sheet1']

    headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    required = [
        'QualificationCode', 'QualificationDescription',
        'SubjectCode', 'SubjectDescription',
        'ModuleCode', 'TopicCode', 'TopicName',
        'AssessmentCriterionCode', 'AssessmentCriterion',
        'LPN', 'LessonPlanTitle', 'BloomLevel',
        'Youtube Video Url', 'Open Source Material for after hour study', 'Open Source 1'
    ]

    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f'Missing columns in workbook: {missing}')

    def get_cell(row_idx, key):
        return ws.cell(row_idx, headers[key]).value

    items = []
    for row_idx in range(2, ws.max_row + 1):
        row = {key: clean(get_cell(row_idx, key)) for key in required}
        if not any(row.values()):
            continue

        youtube_urls = split_urls(row['Youtube Video Url'])
        open_source_urls = split_urls(row['Open Source Material for after hour study'])
        open_source_urls.extend(split_urls(row['Open Source 1']))

        dedup_open = []
        seen_open = set()
        for url in open_source_urls:
            key = url.lower()
            if key in seen_open:
                continue
            seen_open.add(key)
            dedup_open.append(url)

        items.append({
            'excelRow': row_idx,
            'qualificationCode': row['QualificationCode'],
            'qualificationDescription': row['QualificationDescription'],
            'subjectCode': row['SubjectCode'],
            'subjectDescription': row['SubjectDescription'],
            'moduleCode': row['ModuleCode'],
            'topicCode': row['TopicCode'],
            'topicName': row['TopicName'],
            'assessmentCriterionCode': row['AssessmentCriterionCode'],
            'assessmentCriterion': row['AssessmentCriterion'],
            'lpn': row['LPN'],
            'lessonPlanTitle': row['LessonPlanTitle'],
            'bloomLevel': row['BloomLevel'],
            'youtubeUrls': youtube_urls,
            'openSourceUrls': dedup_open,
            'hasYoutube': len(youtube_urls) > 0,
            'hasOpenSource': len(dedup_open) > 0
        })

    payload = {
        'source': XLSX_PATH.name,
        'sheet': ws.title,
        'generatedAtUtc': datetime.now(timezone.utc).isoformat(timespec='seconds'),
        'rowCount': len(items),
        'items': items
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding='utf-8')

    print(f'Wrote {len(items)} rows to {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
